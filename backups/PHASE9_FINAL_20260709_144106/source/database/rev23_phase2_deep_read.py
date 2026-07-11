from pathlib import Path

from openpyxl import load_workbook


MASTER = Path(
    "/Users/test/Desktop/"
    "REDBOX_MASTER_OPERASYON_SISTEMI_"
    "REV23_LOT_SECIMI_RENKLI_NAV.xlsx"
)


def hucre_degeri(value):
    if value is None:
        return ""

    return str(value).strip()


def sayfa_derin_oku(ws):
    print("")
    print("=" * 80)
    print("SAYFA:", ws.title)
    print(
        "BOYUT:",
        ws.max_row,
        "SATIR x",
        ws.max_column,
        "SUTUN",
    )
    print("=" * 80)

    dolu_satir = 0

    for row_no in range(
        1,
        ws.max_row + 1,
    ):
        degerler = []

        for col_no in range(
            1,
            ws.max_column + 1,
        ):
            value = ws.cell(
                row=row_no,
                column=col_no,
            ).value

            if value is not None:
                degerler.append(
                    (
                        ws.cell(
                            row=row_no,
                            column=col_no,
                        ).coordinate,
                        hucre_degeri(value),
                    )
                )

        if not degerler:
            continue

        dolu_satir += 1

        print("")
        print(
            "--- SATIR",
            row_no,
            "---",
        )

        for koordinat, deger in degerler:
            print(
                f"{koordinat}: {deger}"
            )

    print("")
    print(
        "DOLU SATIR SAYISI:",
        dolu_satir,
    )


def main():
    print(
        "=== REV23 PHASE 2 "
        "PAKETLEME + SEVKIYAT "
        "DERIN OKUMA ==="
    )

    if not MASTER.exists():
        raise FileNotFoundError(
            f"MASTER EXCEL BULUNAMADI: {MASTER}"
        )

    print("")
    print("MASTER EXCEL:")
    print(MASTER)

    wb = load_workbook(
        MASTER,
        data_only=False,
        read_only=False,
    )

    print("")
    print("=== WORKBOOK SAYFALARI ===")

    for index, sheet_name in enumerate(
        wb.sheetnames,
        start=1,
    ):
        print(
            index,
            "|",
            sheet_name,
        )

    hedefler = []

    for sheet_name in wb.sheetnames:
        normalized = (
            sheet_name
            .upper()
            .replace("İ", "I")
            .replace("Ş", "S")
            .replace("Ğ", "G")
            .replace("Ü", "U")
            .replace("Ö", "O")
            .replace("Ç", "C")
        )

        if (
            "2026" in normalized
            and (
                "PAKET" in normalized
                or "SEVK" in normalized
            )
        ):
            hedefler.append(
                sheet_name
            )

    print("")
    print("=== PHASE 2 HEDEF SAYFALAR ===")

    if not hedefler:
        print(
            "PAKETLEME / SEVKIYAT "
            "HEDEF SAYFASI BULUNAMADI"
        )

        raise SystemExit(1)

    for sheet_name in hedefler:
        print(
            "HEDEF:",
            sheet_name,
        )

    for sheet_name in hedefler:
        sayfa_derin_oku(
            wb[sheet_name]
        )

    print("")
    print(
        "REV23 PHASE 2 DERIN OKUMA "
        "TAMAMLANDI"
    )

    print("")
    print(
        "NOT: SQL VERISI DEGISTIRILMEDI"
    )


if __name__ == "__main__":
    main()
