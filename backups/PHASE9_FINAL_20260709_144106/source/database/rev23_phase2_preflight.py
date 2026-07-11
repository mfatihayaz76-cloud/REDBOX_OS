from pathlib import Path
from collections import defaultdict

from openpyxl import load_workbook

from database.db import get_connection


MASTER = Path(
    "/Users/test/Desktop/"
    "REDBOX_MASTER_OPERASYON_SISTEMI_"
    "REV23_LOT_SECIMI_RENKLI_NAV.xlsx"
)


def temiz(value):
    if value is None:
        return ""

    return str(value).strip()


def sayi(value):
    if value in (
        None,
        "",
    ):
        return 0.0

    return float(value)


def paketleme_oku(wb):
    ws = wb[
        "22 2026 PAKETLEME"
    ]

    kayitlar = []

    for row_no in range(
        2,
        ws.max_row + 1,
    ):
        uretim_bagi = temiz(
            ws.cell(
                row=row_no,
                column=2,
            ).value
        )

        if not uretim_bagi:
            continue

        net_mamul = ws.cell(
            row=row_no,
            column=3,
        ).value

        if not isinstance(
            net_mamul,
            (int, float),
        ):
            continue

        kayitlar.append({
            "satir": row_no,
            "paketleme_tarihi": temiz(
                ws.cell(
                    row=row_no,
                    column=1,
                ).value
            ),
            "uretim_bagi": uretim_bagi,
            "net_mamul_kg": sayi(
                net_mamul
            ),
            "paket_500": int(
                sayi(
                    ws.cell(
                        row=row_no,
                        column=4,
                    ).value
                )
            ),
            "kg_500": sayi(
                ws.cell(
                    row=row_no,
                    column=5,
                ).value
            ),
            "paket_2500": int(
                sayi(
                    ws.cell(
                        row=row_no,
                        column=6,
                    ).value
                )
            ),
            "kg_2500": sayi(
                ws.cell(
                    row=row_no,
                    column=7,
                ).value
            ),
            "paketleme_firesi": sayi(
                ws.cell(
                    row=row_no,
                    column=8,
                ).value
            ),
            "aciklama": temiz(
                ws.cell(
                    row=row_no,
                    column=11,
                ).value
            ),
        })

    return kayitlar


def sevkiyat_oku(wb):
    ws = wb[
        "23 2026 SEVKİYAT"
    ]

    kayitlar = []

    onceki_tarih = ""

    for row_no in range(
        2,
        ws.max_row + 1,
    ):
        alici = temiz(
            ws.cell(
                row=row_no,
                column=2,
            ).value
        )

        ambalaj = temiz(
            ws.cell(
                row=row_no,
                column=6,
            ).value
        )

        kg_value = ws.cell(
            row=row_no,
            column=8,
        ).value

        if (
            not alici
            or not ambalaj
            or not isinstance(
                kg_value,
                (int, float),
            )
        ):
            continue

        tarih = temiz(
            ws.cell(
                row=row_no,
                column=1,
            ).value
        )

        if tarih:
            onceki_tarih = tarih

        kayitlar.append({
            "satir": row_no,
            "sevk_tarihi": (
                tarih
                or onceki_tarih
            ),
            "alici": alici,
            "lojistik": temiz(
                ws.cell(
                    row=row_no,
                    column=3,
                ).value
            ),
            "plaka": temiz(
                ws.cell(
                    row=row_no,
                    column=4,
                ).value
            ),
            "uretim_bagi": temiz(
                ws.cell(
                    row=row_no,
                    column=5,
                ).value
            ),
            "ambalaj": ambalaj,
            "paket": ws.cell(
                row=row_no,
                column=7,
            ).value,
            "kg": sayi(
                kg_value
            ),
            "not": temiz(
                ws.cell(
                    row=row_no,
                    column=9,
                ).value
            ),
        })

    return kayitlar


def main():
    print(
        "=== REV23 PHASE 2 "
        "PAKETLEME + SEVKIYAT "
        "FINAL PREFLIGHT ==="
    )

    if not MASTER.exists():
        raise FileNotFoundError(
            f"MASTER BULUNAMADI: {MASTER}"
        )

    wb = load_workbook(
        MASTER,
        data_only=False,
        read_only=False,
    )

    paketlemeler = paketleme_oku(
        wb
    )

    sevkiyatlar = sevkiyat_oku(
        wb
    )

    conn = get_connection()

    uretimler = conn.execute("""
        SELECT
            id,
            uretim_tarihi,
            urun_lot_no,
            net_uretim_kg
        FROM uretim
        ORDER BY id
    """).fetchall()

    uretim_map = {
        temiz(
            row["uretim_tarihi"]
        ): row
        for row in uretimler
    }

    print("")
    print(
        "=== PAKETLEME KUTLE "
        "DENKLIGI ==="
    )

    toplam_paketli = 0.0
    toplam_paket_fire = 0.0

    for row in paketlemeler:
        uretim = uretim_map.get(
            row["uretim_bagi"]
        )

        if uretim is None:
            raise ValueError(
                "URETIM BAGI BULUNAMADI: "
                f"{row['uretim_bagi']}"
            )

        paketli_kg = (
            row["kg_500"]
            + row["kg_2500"]
        )

        kapanis = (
            paketli_kg
            + row["paketleme_firesi"]
        )

        fark = round(
            row["net_mamul_kg"]
            - kapanis,
            6,
        )

        durum = (
            "KAPALI"
            if abs(fark) < 0.001
            else "KONTROL"
        )

        print(
            f"{row['paketleme_tarihi']} | "
            f"URETIM: {row['uretim_bagi']} | "
            f"LOT: {uretim['urun_lot_no']} | "
            f"NET: {row['net_mamul_kg']:.3f} | "
            f"PAKETLI: {paketli_kg:.3f} | "
            f"PAKET FIRE: "
            f"{row['paketleme_firesi']:.3f} | "
            f"FARK: {fark:.3f} | "
            f"{durum}"
        )

        if durum != "KAPALI":
            raise ValueError(
                "PAKETLEME KUTLE DENKLIGI "
                "KAPANMADI: "
                f"{row['uretim_bagi']}"
            )

        toplam_paketli += paketli_kg
        toplam_paket_fire += (
            row["paketleme_firesi"]
        )

    print("")
    print(
        "PAKETLEME KAYDI:",
        len(paketlemeler),
    )
    print(
        "TOPLAM PAKETLI MAMUL:",
        f"{toplam_paketli:.3f} KG",
    )
    print(
        "TOPLAM PAKETLEME FIRESI:",
        f"{toplam_paket_fire:.3f} KG",
    )

    print("")
    print(
        "=== SEVKIYAT GERCEK "
        "KAYIT DENETIMI ==="
    )

    toplam_sevk = 0.0
    teyit = []

    for row in sevkiyatlar:
        print(
            f"{row['sevk_tarihi']} | "
            f"{row['alici']} | "
            f"{row['ambalaj']} | "
            f"{row['kg']:.3f} KG | "
            f"URETIM BAGI: "
            f"{row['uretim_bagi']}"
        )

        toplam_sevk += row["kg"]

        if (
            "TEYIT"
            in row[
                "uretim_bagi"
            ].upper()
        ):
            teyit.append(
                row
            )

    print("")
    print(
        "SEVKIYAT KAYDI:",
        len(sevkiyatlar),
    )
    print(
        "TOPLAM SEVK:",
        f"{toplam_sevk:.3f} KG",
    )

    print("")
    print(
        "=== LOT DAGILIMI "
        "TEYIT BEKLEYENLER ==="
    )

    if not teyit:
        print(
            "TEYIT BEKLEYEN "
            "SEVKIYAT YOK"
        )

    for row in teyit:
        print(
            f"{row['sevk_tarihi']} | "
            f"{row['alici']} | "
            f"{row['ambalaj']} | "
            f"{row['kg']:.3f} KG"
        )

    print("")
    print(
        "LOT DAGILIMI TEYIT "
        "KAYDI:",
        len(teyit),
    )

    print("")
    print(
        "=== MEVCUT SQL "
        "PHASE 2 DURUMU ==="
    )

    for tablo in (
        "paketleme",
        "sevkiyat",
        "sevkiyat_kalemleri",
        "mamul_stok_hareketleri",
    ):
        count = conn.execute(
            f"SELECT COUNT(*) FROM {tablo}"
        ).fetchone()[0]

        print(
            f"{tablo:30}: {count}"
        )

    fk = conn.execute(
        "PRAGMA foreign_key_check"
    ).fetchall()

    if fk:
        raise ValueError(
            f"FOREIGN KEY HATASI: {fk}"
        )

    print("")
    print(
        "FOREIGN KEY CHECK: OK"
    )

    conn.close()

    print("")
    print(
        "REV23 PHASE 2 FINAL "
        "PREFLIGHT TAMAMLANDI"
    )

    print("")
    print(
        "NOT: SQL VERISI "
        "DEGISTIRILMEDI"
    )


if __name__ == "__main__":
    main()
