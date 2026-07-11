from pathlib import Path
from datetime import datetime
import shutil
import sqlite3

from openpyxl import load_workbook

from database.db import get_connection
from database.stock_engine import uretim_stok_isle


BASE_DIR = Path(__file__).resolve().parent.parent
DB_PATH = BASE_DIR / "database" / "redbox_os.db"
BACKUP_DIR = BASE_DIR / "backups"

EXCEL_ADAYLARI = [
    Path(
        "/Users/test/Desktop/"
        "REDBOX_MASTER_OPERASYON_SISTEMI_"
        "REV23_LOT_SECIMI_RENKLI_NAV.xlsx"
    ),
    Path(
        "/Users/test/Desktop/"
        "REDBOX_MASTER_OPERASYON_SISTEMI_"
        "REV23_LOT_SECIMI_RENKLI_NAV(1).xlsx"
    ),
]


HAMMADDE_AD_MAP = {
    "Patates unu": "Patates Unu",
    "Mısır nişastası": "Nişasta",
    "Tat mısır unu": "Mısır Unu",
    "Savour tavuk çeşnisi": "Tavuk Çeşnisi",
    "Sarımsak tozu": "Sarımsak Tozu",
    "Karabiber": "Karabiber",
    "Tuz": "Tuz",
    "BENECEL A4M": "Metilselüloz Benecel A4M E461",
}


LOT_KOLONLARI = {
    "Patates Unu": 3,
    "Nişasta": 4,
    "Mısır Unu": 5,
    "Tavuk Çeşnisi": 6,
    "Sarımsak Tozu": 7,
    "Karabiber": 8,
    "Tuz": 9,
    "Metilselüloz Benecel A4M E461": 10,
}


def excel_bul():
    for path in EXCEL_ADAYLARI:
        if path.exists():
            return path

    raise FileNotFoundError(
        "REV23 MASTER EXCEL MASAUSTUNDE BULUNAMADI"
    )


def metin(value):
    if value is None:
        return None

    value = str(value).strip()

    return value or None


def sayi(value):
    if value is None:
        return None

    return float(value)


def tarih(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")

    return str(value).strip()


def uretimleri_oku(wb):
    ws = wb["21 2026 ÜRETİM"]

    kayitlar = []

    for row in range(2, ws.max_row + 1):
        uretim_tarihi = tarih(
            ws.cell(row, 1).value
        )

        parti = ws.cell(row, 2).value

        if not uretim_tarihi or parti in (None, ""):
            continue

        kayit = {
            "uretim_tarihi": uretim_tarihi,
            "parti_sayisi": int(parti),
            "kg_parti": sayi(
                ws.cell(row, 3).value
            ),
            "teorik_uretim_kg": sayi(
                ws.cell(row, 4).value
            ),
            "uretim_firesi_kg": sayi(
                ws.cell(row, 5).value
            ) or 0.0,
            "net_uretim_kg": sayi(
                ws.cell(row, 6).value
            ),
            "urun_lot_no": metin(
                ws.cell(row, 7).value
            ),
            "aciklama": metin(
                ws.cell(row, 8).value
            ),
        }

        if uretim_tarihi == "07.05.2026":
            kayit["uretim_firesi_kg"] = 3.000
            kayit["net_uretim_kg"] = (
                kayit["teorik_uretim_kg"] - 3.000
            )
            kayit["aciklama"] = (
                "Kullanıcı final doğrulaması: "
                "üretim firesi 3.000 kg"
            )

        kayitlar.append(kayit)

    return kayitlar


def lot_baglarini_oku(wb):
    ws = wb["25 2026 LOT BAĞI"]

    lotlar = {}

    for row in range(2, ws.max_row + 1):
        uretim_tarihi = tarih(
            ws.cell(row, 1).value
        )

        urun_lot_no = metin(
            ws.cell(row, 2).value
        )

        if not uretim_tarihi or not urun_lot_no:
            continue

        baglar = {}

        for hammadde, kolon in LOT_KOLONLARI.items():
            lot_no = metin(
                ws.cell(row, kolon).value
            )

            if lot_no:
                baglar[hammadde] = lot_no

        lotlar[uretim_tarihi] = {
            "urun_lot_no": urun_lot_no,
            "baglar": baglar,
        }

    return lotlar


def staging_dogrula(
    uretimler,
    lot_baglari,
):
    print("")
    print(
        "=== REV23 2026 URETIM STAGING DENETIMI ==="
    )

    if len(uretimler) != 11:
        raise ValueError(
            "REV23 URETIM KAYDI 11 OLMALI. "
            f"BULUNAN: {len(uretimler)}"
        )

    gorulen_lotlar = set()

    toplam_parti = 0
    toplam_teorik = 0.0
    toplam_fire = 0.0
    toplam_net = 0.0

    for row in uretimler:
        tarih_key = row["uretim_tarihi"]

        if tarih_key not in lot_baglari:
            raise ValueError(
                "LOT BAGI BULUNAMADI: "
                + tarih_key
            )

        lot_row = lot_baglari[tarih_key]

        excel_lot = lot_row["urun_lot_no"]

        if row["urun_lot_no"] != excel_lot:
            raise ValueError(
                "URETIM LOT / LOT BAGI UYUSMUYOR: "
                f'{tarih_key} | '
                f'{row["urun_lot_no"]} / '
                f'{excel_lot}'
            )

        if row["urun_lot_no"] in gorulen_lotlar:
            raise ValueError(
                "TEKRAR EDEN URUN LOTU: "
                + row["urun_lot_no"]
            )

        gorulen_lotlar.add(
            row["urun_lot_no"]
        )

        eksik_lotlar = [
            hammadde
            for hammadde in LOT_KOLONLARI
            if hammadde not in lot_row["baglar"]
        ]

        if eksik_lotlar:
            raise ValueError(
                f'{tarih_key} LOT BAGI EKSIK: '
                + ", ".join(eksik_lotlar)
            )

        toplam_parti += row["parti_sayisi"]
        toplam_teorik += row["teorik_uretim_kg"]
        toplam_fire += row["uretim_firesi_kg"]
        toplam_net += row["net_uretim_kg"]

        print(
            "OK:",
            tarih_key,
            "| LOT:",
            row["urun_lot_no"],
            "| PARTI:",
            row["parti_sayisi"],
            "| TEORIK:",
            f'{row["teorik_uretim_kg"]:.3f}',
            "| FIRE:",
            f'{row["uretim_firesi_kg"]:.3f}',
            "| NET:",
            f'{row["net_uretim_kg"]:.3f}',
        )

    print("")
    print("URETIM KAYDI:", len(uretimler))
    print("TOPLAM PARTI:", toplam_parti)
    print(
        "TOPLAM TEORIK:",
        f"{toplam_teorik:.3f} KG",
    )
    print(
        "TOPLAM FIRE:",
        f"{toplam_fire:.3f} KG",
    )
    print(
        "TOPLAM NET:",
        f"{toplam_net:.3f} KG",
    )


def mevcut_veri_durumu(conn):
    print("")
    print("=== MEVCUT SQL VERI DURUMU ===")

    tablolar = [
        "depo_kabul",
        "uretim",
        "uretim_recete",
        "uretim_hammadde_lotlari",
        "paketleme",
        "sevkiyat",
        "sevkiyat_kalemleri",
        "mamul_stok_hareketleri",
    ]

    for tablo in tablolar:
        try:
            sayi_row = conn.execute(
                f"SELECT COUNT(*) AS sayi FROM {tablo}"
            ).fetchone()

            print(
                f"{tablo:<30}: "
                f'{sayi_row["sayi"]}'
            )

        except sqlite3.OperationalError:
            print(
                f"{tablo:<30}: TABLO YOK"
            )


def main():
    print(
        "=== REDBOX OS REV23 GERCEK MASTER "
        "2026 AKTARIM HAZIRLIGI ==="
    )

    excel_path = excel_bul()

    print("")
    print("MASTER EXCEL:")
    print(excel_path)

    wb = load_workbook(
        excel_path,
        data_only=False,
    )

    gerekli_sayfalar = [
        "20 2026 DEPO KABUL",
        "21 2026 ÜRETİM",
        "22 2026 PAKETLEME",
        "23 2026 SEVKİYAT",
        "25 2026 LOT BAĞI",
    ]

    for sayfa in gerekli_sayfalar:
        if sayfa not in wb.sheetnames:
            raise ValueError(
                "MASTER SAYFA EKSIK: "
                + sayfa
            )

    if "97 LOT BAĞI ESKİ" in wb.sheetnames:
        print("")
        print(
            "OK: 97 LOT BAGI ESKI "
            "AKTARIM DISI"
        )

    uretimler = uretimleri_oku(wb)
    lot_baglari = lot_baglarini_oku(wb)

    staging_dogrula(
        uretimler,
        lot_baglari,
    )

    conn = get_connection()

    try:
        mevcut_veri_durumu(conn)

        fk = conn.execute(
            "PRAGMA foreign_key_check"
        ).fetchall()

        if fk:
            raise ValueError(
                "FOREIGN KEY CHECK HATALI"
            )

        print("")
        print("FOREIGN KEY CHECK: OK")

    finally:
        conn.close()

    print("")
    print(
        "REV23 MASTER OKUMA VE "
        "11 URETIM STAGING DOGRULAMASI BASARILI"
    )

    print("")
    print(
        "NOT: BU ASAMADA SQL VERISI SILINMEDI "
        "VE DEGISTIRILMEDI"
    )


if __name__ == "__main__":
    main()
