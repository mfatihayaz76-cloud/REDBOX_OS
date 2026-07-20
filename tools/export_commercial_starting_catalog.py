import argparse
import csv
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parent.parent

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from database.recipe_import_engine import (
    ALL_HEADERS,
    CATALOG_SHEET_NAME,
)
from tools.create_recipe_catalog_template import (
    create_xlsx_template,
)


def _read_authorized_catalog(database_path):
    connection = sqlite3.connect(
        f"file:{database_path}?mode=ro",
        uri=True,
    )
    connection.row_factory = sqlite3.Row

    try:
        rows = connection.execute(
            """
            SELECT
                u.urun_kodu,
                u.urun_adi,
                r.recete_kodu,
                r.ad AS recete_adi,
                r.revizyon_no,
                r.gecerlilik_tarihi,
                r.durum,
                r.parti_teorik_kg,
                r.proses_suyu_kg,
                h.ad AS hammadde_adi,
                rk.miktar_kg,
                u.kategori,
                u.barkod,
                u.birim,
                u.raf_omru_gun,
                u.saklama_sicakligi,
                u.aciklama AS urun_aciklama,
                r.revizyon_aciklamasi
            FROM receteler r
            JOIN urunler u
              ON u.id = r.urun_id
            JOIN recete_kalemleri rk
              ON rk.recete_id = r.id
            JOIN hammaddeler h
              ON h.id = rk.hammadde_id
            WHERE r.durum IN ('AKTIF', 'ARSIV')
            ORDER BY
                u.urun_kodu,
                r.id,
                rk.id
            """
        ).fetchall()

        materials = [
            row["ad"]
            for row in connection.execute(
                """
                SELECT ad
                FROM hammaddeler
                WHERE aktif = 1
                ORDER BY ad COLLATE NOCASE
                """
            ).fetchall()
        ]

        integrity = connection.execute(
            "PRAGMA integrity_check"
        ).fetchone()[0]

    finally:
        connection.close()

    if integrity != "ok":
        raise RuntimeError(
            "Kaynak veritabanı bütünlük kontrolü başarısız."
        )

    if not rows:
        raise RuntimeError(
            "Yetkili AKTIF/ARSIV reçete verisi bulunamadı."
        )

    normalized = []

    for row in rows:
        record = {
            header: (
                ""
                if row[header] is None
                else row[header]
            )
            for header in ALL_HEADERS
        }

        record["parti_teorik_kg"] = float(
            record["parti_teorik_kg"]
        )
        record["proses_suyu_kg"] = float(
            record["proses_suyu_kg"]
        )
        record["miktar_kg"] = float(
            record["miktar_kg"]
        )

        if record["raf_omru_gun"] != "":
            record["raf_omru_gun"] = int(
                record["raf_omru_gun"]
            )

        normalized.append(record)

    return normalized, materials


def _write_csv(output_path, rows):
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with output_path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=ALL_HEADERS,
            delimiter=";",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(output_path, rows, materials):
    create_xlsx_template(
        output_path,
        materials,
    )

    workbook = load_workbook(output_path)
    catalog = workbook[CATALOG_SHEET_NAME]

    column_widths = {
        "urun_kodu": 14,
        "urun_adi": 22,
        "recete_kodu": 18,
        "recete_adi": 42,
        "revizyon_no": 13,
        "gecerlilik_tarihi": 18,
        "durum": 14,
        "parti_teorik_kg": 19,
        "proses_suyu_kg": 19,
        "hammadde_adi": 38,
        "miktar_kg": 14,
        "kategori": 30,
        "barkod": 18,
        "birim": 12,
        "raf_omru_gun": 17,
        "saklama_sicakligi": 23,
        "urun_aciklama": 48,
        "revizyon_aciklamasi": 54,
    }
    wrapped_headers = {
        "recete_adi",
        "hammadde_adi",
        "kategori",
        "urun_aciklama",
        "revizyon_aciklamasi",
    }
    numeric_formats = {
        "parti_teorik_kg": "0.000",
        "proses_suyu_kg": "0.000",
        "miktar_kg": "0.000",
        "raf_omru_gun": "0",
    }

    for column, header in enumerate(
        ALL_HEADERS,
        start=1,
    ):
        header_cell = catalog.cell(
            row=1,
            column=column,
        )
        catalog.column_dimensions[
            header_cell.column_letter
        ].width = column_widths[header]
        header_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    catalog.row_dimensions[1].height = 48

    even_fill = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )

    for row_number, record in enumerate(
        rows,
        start=2,
    ):
        catalog.append([
            record[header]
            for header in ALL_HEADERS
        ])

        for column, header in enumerate(
            ALL_HEADERS,
            start=1,
        ):
            cell = catalog.cell(
                row=row_number,
                column=column,
            )
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=header in wrapped_headers,
            )

            if header in numeric_formats:
                cell.number_format = numeric_formats[header]

            if header == "revizyon_no":
                cell.number_format = "@"

            if row_number % 2 == 0:
                cell.fill = even_fill

        catalog.row_dimensions[row_number].height = 32

    workbook.save(output_path)
    workbook.close()


def export_catalog(
    database_path,
    output_directory,
):
    database_path = Path(database_path).resolve()
    output_directory = Path(
        output_directory
    ).resolve()

    rows, materials = _read_authorized_catalog(
        database_path
    )

    csv_path = (
        output_directory
        / "REDBOX_OS_TICARI_BASLANGIC_KATALOGU.csv"
    )
    xlsx_path = (
        output_directory
        / "REDBOX_OS_TICARI_BASLANGIC_KATALOGU.xlsx"
    )

    _write_csv(csv_path, rows)
    _write_xlsx(xlsx_path, rows, materials)

    recipe_keys = {
        (
            row["urun_kodu"],
            row["recete_kodu"],
            row["revizyon_no"],
        )
        for row in rows
    }
    product_codes = {
        row["urun_kodu"]
        for row in rows
    }

    return {
        "csv_path": csv_path,
        "xlsx_path": xlsx_path,
        "satir_sayisi": len(rows),
        "urun_sayisi": len(product_codes),
        "recete_sayisi": len(recipe_keys),
        "hammadde_referansi": len(materials),
    }


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Canlı kontrollü katalogdan gerçek ticari "
            "başlangıç CSV/XLSX dosyalarını üretir."
        )
    )
    parser.add_argument(
        "--database",
        type=Path,
        default=(
            PROJECT_ROOT
            / "database"
            / "redbox_os.db"
        ),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=PROJECT_ROOT / "catalogs",
    )
    args = parser.parse_args()

    result = export_catalog(
        args.database,
        args.output_dir,
    )

    print(
        f"CSV: {result['csv_path']}"
    )
    print(
        f"XLSX: {result['xlsx_path']}"
    )
    print(
        f"ÜRÜN: {result['urun_sayisi']}"
    )
    print(
        f"REÇETE/REVİZYON: {result['recete_sayisi']}"
    )
    print(
        f"KATALOG SATIRI: {result['satir_sayisi']}"
    )
    print(
        "HAMMADDE REFERANSI: "
        f"{result['hammadde_referansi']}"
    )


if __name__ == "__main__":
    main()
