import argparse
import csv
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import (
    DataValidation,
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from database.recipe_import_engine import (
    ALL_HEADERS,
    CATALOG_SHEET_NAME,
    REQUIRED_HEADERS,
)


FIELD_DESCRIPTIONS = {
    "urun_kodu": (
        "Evet",
        "Benzersiz ürün kodu; harf/rakam/tire.",
        "URN001",
    ),
    "urun_adi": (
        "Evet",
        "Benzersiz ticari ürün adı.",
        "ÖRNEK ÜRÜN",
    ),
    "kategori": (
        "Hayır",
        "Ürün kategorisi.",
        "Toz Ürün",
    ),
    "barkod": (
        "Hayır",
        "Varsa benzersiz barkod.",
        "",
    ),
    "birim": (
        "Hayır",
        "Ürün birimi; mevcut sözleşmede KG.",
        "KG",
    ),
    "raf_omru_gun": (
        "Hayır",
        "Raf ömrü, pozitif tam gün.",
        "365",
    ),
    "saklama_sicakligi": (
        "Hayır",
        "Saklama sıcaklığı/metni.",
        "-18°C",
    ),
    "urun_aciklama": (
        "Hayır",
        "Ürün kartı açıklaması.",
        "",
    ),
    "recete_kodu": (
        "Evet",
        "Ürün içindeki reçete kodu.",
        "URN001-REC",
    ),
    "recete_adi": (
        "Evet",
        "Benzersiz reçete/revizyon adı.",
        "ÖRNEK ÜRÜN ANA REÇETE REV 00",
    ),
    "revizyon_no": (
        "Evet",
        "Revizyon numarası.",
        "00",
    ),
    "gecerlilik_tarihi": (
        "Evet*",
        "ONAYLI/AKTIF için GG.AA.YYYY.",
        "19.07.2026",
    ),
    "durum": (
        "Evet",
        (
            "TASLAK, INCELEMEDE, ONAYLI, "
            "AKTIF, PASIF veya ARSIV."
        ),
        "TASLAK",
    ),
    "parti_teorik_kg": (
        "Evet",
        "Bir partinin toplam teorik kg değeri.",
        "20.000",
    ),
    "proses_suyu_kg": (
        "Evet",
        (
            "Kütleye dahil, stoka dahil olmayan su; "
            "kullanılmıyorsa 0."
        ),
        "0.000",
    ),
    "revizyon_aciklamasi": (
        "Hayır",
        "Revizyonun kontrollü değişiklik özeti.",
        "",
    ),
    "hammadde_adi": (
        "Evet",
        "Sistemdeki aktif hammadde adıyla birebir.",
        "Hammadde adı",
    ),
    "miktar_kg": (
        "Evet",
        "Bu reçete kaleminin pozitif kg miktarı.",
        "1.000",
    ),
}


def active_materials(database_path):
    conn = sqlite3.connect(database_path)

    try:
        return [
            row[0]
            for row in conn.execute(
                """
                SELECT ad
                FROM hammaddeler
                WHERE aktif = 1
                ORDER BY ad COLLATE NOCASE
                """
            ).fetchall()
        ]
    finally:
        conn.close()


def create_csv_template(output_path):
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with output_path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as handle:
        writer = csv.writer(
            handle,
            delimiter=";",
            lineterminator="\n",
        )
        writer.writerow(ALL_HEADERS)


def create_xlsx_template(
    output_path,
    materials,
):
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()
    catalog = workbook.active
    catalog.title = CATALOG_SHEET_NAME
    catalog.freeze_panes = "A2"
    catalog.auto_filter.ref = (
        f"A1:R5000"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )
    required_fill = PatternFill(
        "solid",
        fgColor="C65911",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for column, header in enumerate(
        ALL_HEADERS,
        start=1,
    ):
        cell = catalog.cell(
            row=1,
            column=column,
            value=header,
        )
        cell.font = header_font
        cell.fill = (
            required_fill
            if header in REQUIRED_HEADERS
            else header_fill
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        catalog.column_dimensions[
            cell.column_letter
        ].width = max(15, min(28, len(header) + 4))

    catalog.row_dimensions[1].height = 36

    status_validation = DataValidation(
        type="list",
        formula1=(
            '"TASLAK,INCELEMEDE,ONAYLI,'
            'AKTIF,PASIF,ARSIV"'
        ),
        allow_blank=False,
    )
    unit_validation = DataValidation(
        type="list",
        formula1='"KG"',
        allow_blank=True,
    )
    catalog.add_data_validation(status_validation)
    catalog.add_data_validation(unit_validation)

    status_column = ALL_HEADERS.index("durum") + 1
    unit_column = ALL_HEADERS.index("birim") + 1

    status_validation.add(
        f"{catalog.cell(2, status_column).column_letter}"
        f"2:{catalog.cell(5000, status_column).column_letter}"
        "5000"
    )
    unit_validation.add(
        f"{catalog.cell(2, unit_column).column_letter}"
        f"2:{catalog.cell(5000, unit_column).column_letter}"
        "5000"
    )

    instructions = workbook.create_sheet(
        "ALAN_ACIKLAMALARI"
    )
    instructions.append([
        "alan",
        "zorunlu",
        "aciklama",
        "bicim_ornegi",
    ])

    for cell in instructions[1]:
        cell.font = header_font
        cell.fill = header_fill

    for header in ALL_HEADERS:
        required, description, example = (
            FIELD_DESCRIPTIONS[header]
        )
        instructions.append([
            header,
            required,
            description,
            example,
        ])

    instructions.freeze_panes = "A2"
    instructions.column_dimensions["A"].width = 28
    instructions.column_dimensions["B"].width = 12
    instructions.column_dimensions["C"].width = 70
    instructions.column_dimensions["D"].width = 38

    reference = workbook.create_sheet(
        "HAMMADDE_REFERANSI"
    )
    reference.append([
        "aktif_hammadde_adi",
        "kullanim_notu",
    ])

    for cell in reference[1]:
        cell.font = header_font
        cell.fill = header_fill

    for material in materials:
        reference.append([
            material,
            (
                "RECETE_KATALOGU sayfasında "
                "birebir bu ad kullanılmalıdır."
            ),
        ])

    reference.freeze_panes = "A2"
    reference.column_dimensions["A"].width = 48
    reference.column_dimensions["B"].width = 68

    workbook.save(output_path)
    workbook.close()


def main():
    parser = argparse.ArgumentParser(
        description=(
            "REDBOX OS reçete katalog CSV/XLSX "
            "şablonlarını oluşturur."
        )
    )
    parser.add_argument(
        "--database",
        type=Path,
        default=(
            PROJECT_ROOT
            / "database"
            / "redbox_os.db"
        ),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=(
            PROJECT_ROOT
            / "templates"
        ),
    )
    args = parser.parse_args()

    materials = active_materials(args.database)
    xlsx_path = (
        args.output_dir
        / "REDBOX_OS_RECETE_KATALOGU_SABLONU.xlsx"
    )
    csv_path = (
        args.output_dir
        / "REDBOX_OS_RECETE_KATALOGU_SABLONU.csv"
    )

    create_xlsx_template(
        xlsx_path,
        materials,
    )
    create_csv_template(csv_path)

    print(f"XLSX: {xlsx_path}")
    print(f"CSV : {csv_path}")
    print(f"AKTIF HAMMADDE: {len(materials)}")


if __name__ == "__main__":
    main()
