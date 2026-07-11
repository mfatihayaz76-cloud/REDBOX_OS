from collections import defaultdict

from database.db import get_connection
from database.import_2026_rev23 import (
    excel_bul,
    uretimleri_oku,
    lot_baglarini_oku,
)

from openpyxl import load_workbook


HAMMADDE_AD_MAP = {
    "Patates unu": "Patates Unu",
    "Mısır nişastası": "Nişasta",
    "Tat mısır unu": "Mısır Unu",
    "Savour tavuk çeşnisi": "Tavuk Çeşnisi",
    "Sarımsak tozu": "Sarımsak Tozu",
    "Karabiber": "Karabiber",
    "Tuz": "Tuz",
    "BENECEL A4M": "Metilselüloz Benecel A4M E461",
}


def aktif_recete_oku(conn):
    recete = conn.execute("""
        SELECT
            id,
            ad,
            parti_teorik_kg
        FROM receteler
        WHERE aktif = 1
    """).fetchone()

    if recete is None:
        raise ValueError(
            "AKTIF RECETE BULUNAMADI"
        )

    kalemler = conn.execute("""
        SELECT
            h.ad AS hammadde,
            rk.miktar_kg
        FROM recete_kalemleri rk
        JOIN hammaddeler h
          ON h.id = rk.hammadde_id
        WHERE rk.recete_id = ?
        ORDER BY h.id
    """, (
        recete["id"],
    )).fetchall()

    return recete, kalemler


def excel_depo_lotlarini_oku(wb):
    ws = wb["20 2026 DEPO KABUL"]

    lotlar = {}

    for row in range(2, ws.max_row + 1):
        excel_hammadde = ws.cell(
            row, 2
        ).value

        lot_no = ws.cell(
            row, 6
        ).value

        if not excel_hammadde or not lot_no:
            continue

        excel_hammadde = str(
            excel_hammadde
        ).strip()

        if excel_hammadde not in HAMMADDE_AD_MAP:
            continue

        hammadde = HAMMADDE_AD_MAP[
            excel_hammadde
        ]

        key = (
            hammadde,
            str(lot_no).strip(),
        )

        lotlar[key] = {
            "hammadde": hammadde,
            "lot_no": str(lot_no).strip(),
            "tedarikci": (
                str(ws.cell(row, 3).value).strip()
                if ws.cell(row, 3).value
                else None
            ),
            "kabul_tarihi": (
                str(ws.cell(row, 1).value).strip()
                if ws.cell(row, 1).value
                else None
            ),
            "uretim_tarihi": (
                str(ws.cell(row, 7).value).strip()
                if ws.cell(row, 7).value
                else None
            ),
            "skt_tett": (
                str(ws.cell(row, 8).value).strip()
                if ws.cell(row, 8).value
                else None
            ),
            "kabul_durumu": (
                str(ws.cell(row, 9).value).strip()
                if ws.cell(row, 9).value
                else None
            ),
            "aciklama": (
                str(ws.cell(row, 10).value).strip()
                if ws.cell(row, 10).value
                else None
            ),
        }

    return lotlar


def main():
    print(
        "=== REV23 PHASE 1 FINAL PREFLIGHT ==="
    )

    excel_path = excel_bul()

    wb = load_workbook(
        excel_path,
        data_only=False,
    )

    uretimler = uretimleri_oku(wb)
    lot_baglari = lot_baglarini_oku(wb)
    depo_lotlari = excel_depo_lotlarini_oku(wb)

    conn = get_connection()

    try:
        recete, kalemler = aktif_recete_oku(
            conn
        )

        toplam_parti = sum(
            row["parti_sayisi"]
            for row in uretimler
        )

        print("")
        print("=== AKTIF RECETE ===")
        print(
            "RECETE:",
            recete["ad"],
        )
        print(
            "PARTI TEORIK:",
            f'{float(recete["parti_teorik_kg"]):.3f} KG',
        )

        print("")
        print("=== 2026 URETIM KAPSAMI ===")
        print(
            "URETIM KAYDI:",
            len(uretimler),
        )
        print(
            "TOPLAM PARTI:",
            toplam_parti,
        )

        print("")
        print(
            "=== 121 PARTI HAMMADDE IHTIYACI ==="
        )

        toplam_ihtiyac = {}

        for kalem in kalemler:
            miktar = (
                float(kalem["miktar_kg"])
                * toplam_parti
            )

            toplam_ihtiyac[
                kalem["hammadde"]
            ] = miktar

            print(
                f'{kalem["hammadde"]:<45}'
                f'{miktar:>12.3f} kg'
            )

        print("")
        print(
            "=== REV23 URETIM LOT BAGI "
            "KULLANIM DAGILIMI ==="
        )

        lot_ihtiyaclari = defaultdict(float)
        eksik_baglar = []

        recete_map = {
            row["hammadde"]:
            float(row["miktar_kg"])
            for row in kalemler
        }

        for uretim in uretimler:
            tarih = uretim["uretim_tarihi"]
            parti = uretim["parti_sayisi"]

            bag = lot_baglari[tarih]["baglar"]

            print("")
            print(
                tarih,
                "|",
                uretim["urun_lot_no"],
                "|",
                parti,
                "PARTI",
            )

            for hammadde, parti_kg in recete_map.items():
                ihtiyac = parti_kg * parti

                lot_no = bag.get(
                    hammadde
                )

                if not lot_no:
                    eksik_baglar.append({
                        "tarih": tarih,
                        "urun_lot": uretim[
                            "urun_lot_no"
                        ],
                        "hammadde": hammadde,
                        "ihtiyac": ihtiyac,
                    })

                    print(
                        "  EKSIK LOT:",
                        hammadde,
                        "| IHTIYAC:",
                        f"{ihtiyac:.3f} kg",
                    )

                    continue

                lot_ihtiyaclari[
                    (
                        hammadde,
                        lot_no,
                    )
                ] += ihtiyac

                print(
                    " ",
                    hammadde,
                    "| LOT:",
                    lot_no,
                    "|",
                    f"{ihtiyac:.3f} kg",
                )

        print("")
        print(
            "=== LOT BAZLI GEREKLI "
            "DEPO GIRIS MIKTARLARI ==="
        )

        eksik_depo_lotlari = []

        for key, miktar in sorted(
            lot_ihtiyaclari.items()
        ):
            hammadde, lot_no = key

            depo = depo_lotlari.get(
                key
            )

            durum = (
                "REV23 DEPO LOTU VAR"
                if depo
                else "DEPO LOT KARTI YOK"
            )

            print(
                f'{hammadde:<42}'
                f' | {lot_no:<18}'
                f' | {miktar:>9.3f} kg'
                f' | {durum}'
            )

            if depo is None:
                eksik_depo_lotlari.append(
                    key
                )

        print("")
        print(
            "=== TARIHSEL EKSIK LOT BAGLARI ==="
        )

        if not eksik_baglar:
            print(
                "TARIHSEL EKSIK LOT BAGI YOK"
            )
        else:
            for row in eksik_baglar:
                print(
                    row["tarih"],
                    "| URUN LOT:",
                    row["urun_lot"],
                    "|",
                    row["hammadde"],
                    "|",
                    f'{row["ihtiyac"]:.3f} kg',
                )

        print("")
        print(
            "TARIHSEL EKSIK LOT BAGI SAYISI:",
            len(eksik_baglar),
        )

        print(
            "DEPO LOT KARTI EKSIK SAYISI:",
            len(eksik_depo_lotlari),
        )

        print("")
        print(
            "=== MEVCUT SQL KAPSAMI ==="
        )

        for tablo in [
            "depo_kabul",
            "uretim",
            "uretim_recete",
            "uretim_hammadde_lotlari",
        ]:
            row = conn.execute(
                f"""
                SELECT COUNT(*) AS sayi
                FROM {tablo}
                """
            ).fetchone()

            print(
                f"{tablo:<30}: "
                f'{row["sayi"]}'
            )

        fk = conn.execute(
            "PRAGMA foreign_key_check"
        ).fetchall()

        if fk:
            raise ValueError(
                "FOREIGN KEY CHECK HATALI"
            )

        print("")
        print(
            "FOREIGN KEY CHECK: OK"
        )

    finally:
        conn.close()

    print("")
    print(
        "REV23 PHASE 1 FINAL PREFLIGHT "
        "TAMAMLANDI"
    )

    print("")
    print(
        "NOT: SQL VERISI DEGISTIRILMEDI"
    )


if __name__ == "__main__":
    main()
