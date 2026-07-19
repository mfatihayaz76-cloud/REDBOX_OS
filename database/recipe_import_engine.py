import csv
import hashlib
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from database.audit_engine import denetim_kaydi_ekle


CATALOG_SHEET_NAME = "RECETE_KATALOGU"

REQUIRED_HEADERS = (
    "urun_kodu",
    "urun_adi",
    "recete_kodu",
    "recete_adi",
    "revizyon_no",
    "gecerlilik_tarihi",
    "durum",
    "parti_teorik_kg",
    "proses_suyu_kg",
    "hammadde_adi",
    "miktar_kg",
)

OPTIONAL_HEADERS = (
    "kategori",
    "barkod",
    "birim",
    "raf_omru_gun",
    "saklama_sicakligi",
    "urun_aciklama",
    "revizyon_aciklamasi",
)

ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS

ALLOWED_RECIPE_STATUSES = {
    "TASLAK",
    "INCELEMEDE",
    "ONAYLI",
    "AKTIF",
    "PASIF",
    "ARSIV",
}

CODE_PATTERN = re.compile(
    r"^[A-Z0-9][A-Z0-9._-]{1,49}$"
)

REVISION_PATTERN = re.compile(r"^[0-9]{2}$")

MASS_TOLERANCE_KG = 0.000001


def _text(value):
    if value is None:
        return ""

    return " ".join(str(value).strip().split())


def _key(value):
    return _text(value).casefold()


def _code(value):
    return _text(value).upper()


def _error(errors, row, field, code, message):
    errors.append({
        "satir": row,
        "alan": field,
        "kod": code,
        "mesaj": message,
    })


def _parse_float(
    value,
    row,
    field,
    errors,
    minimum=None,
    strictly_positive=False,
):
    text = _text(value).replace(",", ".")

    if not text:
        _error(
            errors,
            row,
            field,
            "ZORUNLU_DEGER",
            f"{field} boş olamaz.",
        )
        return None

    try:
        number = float(text)
    except ValueError:
        _error(
            errors,
            row,
            field,
            "GECERSIZ_SAYI",
            f"{field} sayısal olmalıdır.",
        )
        return None

    if strictly_positive and number <= 0:
        _error(
            errors,
            row,
            field,
            "POZITIF_OLMALI",
            f"{field} sıfırdan büyük olmalıdır.",
        )
        return None

    if minimum is not None and number < minimum:
        _error(
            errors,
            row,
            field,
            "ALT_SINIR",
            f"{field} en az {minimum} olmalıdır.",
        )
        return None

    return number


def _parse_integer(
    value,
    row,
    field,
    errors,
    minimum=None,
):
    text = _text(value)

    if not text:
        return None

    try:
        number = int(text)
    except ValueError:
        _error(
            errors,
            row,
            field,
            "GECERSIZ_TAMSAYI",
            f"{field} tam sayı olmalıdır.",
        )
        return None

    if minimum is not None and number < minimum:
        _error(
            errors,
            row,
            field,
            "ALT_SINIR",
            f"{field} en az {minimum} olmalıdır.",
        )
        return None

    return number


def _validate_date(value, row, field, errors):
    text = _text(value)

    if not text:
        return ""

    try:
        datetime.strptime(text, "%d.%m.%Y")
    except ValueError:
        _error(
            errors,
            row,
            field,
            "GECERSIZ_TARIH",
            f"{field} GG.AA.YYYY biçiminde olmalıdır.",
        )

    return text


def _normalise_header(value):
    return _text(value).lower()


def _rows_from_csv(path):
    with path.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as handle:
        sample = handle.read(4096)
        handle.seek(0)

        try:
            dialect = csv.Sniffer().sniff(
                sample,
                delimiters=",;\t",
            )
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(
            handle,
            dialect=dialect,
        )

        if reader.fieldnames is None:
            return [], []

        headers = [
            _normalise_header(header)
            for header in reader.fieldnames
        ]

        rows = []

        for row_number, raw in enumerate(
            reader,
            start=2,
        ):
            row = {
                _normalise_header(key): value
                for key, value in raw.items()
                if key is not None
            }

            if any(_text(value) for value in row.values()):
                row["_satir"] = row_number
                rows.append(row)

        return headers, rows


def _rows_from_xlsx(path):
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    try:
        if CATALOG_SHEET_NAME in workbook.sheetnames:
            sheet = workbook[CATALOG_SHEET_NAME]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        iterator = sheet.iter_rows(values_only=True)
        first_row = next(iterator, None)

        if first_row is None:
            return [], []

        headers = [
            _normalise_header(value)
            for value in first_row
        ]
        rows = []

        for row_number, values in enumerate(
            iterator,
            start=2,
        ):
            row = {
                header: (
                    values[index]
                    if index < len(values)
                    else None
                )
                for index, header in enumerate(headers)
                if header
            }

            if any(_text(value) for value in row.values()):
                row["_satir"] = row_number
                rows.append(row)

        return headers, rows
    finally:
        workbook.close()


def katalog_satirlarini_oku(file_path):
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(
            f"Katalog dosyası bulunamadı: {path}"
        )

    suffix = path.suffix.lower()

    if suffix == ".csv":
        return _rows_from_csv(path)

    if suffix == ".xlsx":
        return _rows_from_xlsx(path)

    raise ValueError(
        "Yalnızca .csv ve .xlsx katalog dosyaları "
        "desteklenir."
    )


def _normalise_row(raw, errors):
    row_number = int(raw.get("_satir", 0))

    required_text_fields = (
        "urun_kodu",
        "urun_adi",
        "recete_kodu",
        "recete_adi",
        "revizyon_no",
        "durum",
        "hammadde_adi",
    )

    for field in required_text_fields:
        if not _text(raw.get(field)):
            _error(
                errors,
                row_number,
                field,
                "ZORUNLU_DEGER",
                f"{field} boş olamaz.",
            )

    product_code = _code(raw.get("urun_kodu"))
    recipe_code = _code(raw.get("recete_kodu"))
    status = _code(raw.get("durum"))

    for field, code_value in (
        ("urun_kodu", product_code),
        ("recete_kodu", recipe_code),
    ):
        if (
            code_value
            and not CODE_PATTERN.fullmatch(code_value)
        ):
            _error(
                errors,
                row_number,
                field,
                "GECERSIZ_KOD",
                (
                    f"{field}; harf, rakam, nokta, "
                    "alt çizgi veya tire içermelidir."
                ),
            )

    revision = _text(raw.get("revizyon_no"))

    if revision and not REVISION_PATTERN.fullmatch(
        revision
    ):
        _error(
            errors,
            row_number,
            "revizyon_no",
            "GECERSIZ_REVIZYON",
            (
                "Revizyon numarası iki rakamdan "
                "oluşmalıdır; örnek: 00."
            ),
        )

    if status and status not in ALLOWED_RECIPE_STATUSES:
        _error(
            errors,
            row_number,
            "durum",
            "GECERSIZ_DURUM",
            f"Geçersiz reçete durumu: {status}",
        )

    validity_date = _validate_date(
        raw.get("gecerlilik_tarihi"),
        row_number,
        "gecerlilik_tarihi",
        errors,
    )

    if status in {"ONAYLI", "AKTIF"} and not validity_date:
        _error(
            errors,
            row_number,
            "gecerlilik_tarihi",
            "AKTIF_TARIH_ZORUNLU",
            (
                "ONAYLI veya AKTIF reçetede "
                "geçerlilik tarihi zorunludur."
            ),
        )

    theoretical_kg = _parse_float(
        raw.get("parti_teorik_kg"),
        row_number,
        "parti_teorik_kg",
        errors,
        strictly_positive=True,
    )
    process_water_kg = _parse_float(
        raw.get("proses_suyu_kg"),
        row_number,
        "proses_suyu_kg",
        errors,
        minimum=0,
    )
    material_kg = _parse_float(
        raw.get("miktar_kg"),
        row_number,
        "miktar_kg",
        errors,
        strictly_positive=True,
    )
    shelf_life_days = _parse_integer(
        raw.get("raf_omru_gun"),
        row_number,
        "raf_omru_gun",
        errors,
        minimum=1,
    )

    unit = _code(raw.get("birim")) or "KG"

    if unit != "KG":
        _error(
            errors,
            row_number,
            "birim",
            "GECERSIZ_BIRIM",
            "Ürün birimi KG olmalıdır.",
        )

    return {
        "satir": row_number,
        "urun_kodu": product_code,
        "urun_adi": _text(raw.get("urun_adi")),
        "kategori": _text(raw.get("kategori")),
        "barkod": _text(raw.get("barkod")),
        "birim": unit,
        "raf_omru_gun": shelf_life_days,
        "saklama_sicakligi": _text(
            raw.get("saklama_sicakligi")
        ),
        "urun_aciklama": _text(
            raw.get("urun_aciklama")
        ),
        "recete_kodu": recipe_code,
        "recete_adi": _text(raw.get("recete_adi")),
        "revizyon_no": revision,
        "gecerlilik_tarihi": validity_date,
        "durum": status,
        "parti_teorik_kg": theoretical_kg,
        "proses_suyu_kg": process_water_kg,
        "revizyon_aciklamasi": _text(
            raw.get("revizyon_aciklamasi")
        ),
        "hammadde_adi": _text(
            raw.get("hammadde_adi")
        ),
        "miktar_kg": material_kg,
    }


def katalog_on_kontrol(conn, headers, raw_rows):
    errors = []
    warnings = []

    normalised_headers = {
        _normalise_header(header)
        for header in headers
        if _normalise_header(header)
    }
    missing_headers = [
        header
        for header in REQUIRED_HEADERS
        if header not in normalised_headers
    ]

    for header in missing_headers:
        _error(
            errors,
            1,
            header,
            "EKSIK_KOLON",
            f"Zorunlu kolon eksik: {header}",
        )

    unknown_headers = sorted(
        normalised_headers - set(ALL_HEADERS)
    )

    for header in unknown_headers:
        warnings.append({
            "satir": 1,
            "alan": header,
            "kod": "BILINMEYEN_KOLON",
            "mesaj": (
                f"Bilinmeyen kolon yok sayılacak: {header}"
            ),
        })

    rows = [
        _normalise_row(raw, errors)
        for raw in raw_rows
    ]

    if not rows:
        _error(
            errors,
            1,
            "",
            "BOS_KATALOG",
            "Katalogda veri satırı bulunamadı.",
        )

    material_rows = conn.execute("""
        SELECT id, ad, birim, aktif
        FROM hammaddeler
    """).fetchall()
    materials = {
        _key(row["ad"]): {
            "id": int(row["id"]),
            "ad": row["ad"],
            "birim": row["birim"],
            "aktif": int(row["aktif"]),
        }
        for row in material_rows
    }

    existing_products = conn.execute("""
        SELECT
            id,
            urun_kodu,
            urun_adi,
            barkod,
            aktif
        FROM urunler
    """).fetchall()
    products_by_code = {
        _key(row["urun_kodu"]): row
        for row in existing_products
    }
    products_by_name = {
        _key(row["urun_adi"]): row
        for row in existing_products
    }
    products_by_barcode = {
        _key(row["barkod"]): row
        for row in existing_products
        if _text(row["barkod"])
    }
    existing_recipe_names = {
        _key(row["ad"])
        for row in conn.execute(
            "SELECT ad FROM receteler"
        ).fetchall()
    }

    existing_recipes = {
        (
            _key(row["urun_kodu"]),
            _key(row["recete_kodu"]),
            _key(row["revizyon_no"]),
        )
        for row in conn.execute("""
            SELECT
                u.urun_kodu,
                r.recete_kodu,
                r.revizyon_no
            FROM receteler AS r
            JOIN urunler AS u
              ON u.id = r.urun_id
        """).fetchall()
    }

    existing_active_products = {
        _key(row["urun_kodu"])
        for row in conn.execute("""
            SELECT u.urun_kodu
            FROM receteler AS r
            JOIN urunler AS u
              ON u.id = r.urun_id
            WHERE r.aktif = 1
        """).fetchall()
    }

    products = {}
    recipes = {}
    imported_product_names = {}
    imported_barcodes = {}
    imported_recipe_names = {}
    imported_active_products = set()

    for row in rows:
        row_number = row["satir"]
        product_code_key = _key(row["urun_kodu"])
        product_name_key = _key(row["urun_adi"])
        recipe_key = (
            product_code_key,
            _key(row["recete_kodu"]),
            _key(row["revizyon_no"]),
        )

        product_signature = (
            row["urun_adi"],
            row["kategori"],
            row["barkod"],
            row["birim"],
            row["raf_omru_gun"],
            row["saklama_sicakligi"],
            row["urun_aciklama"],
        )

        if product_code_key in products:
            if (
                products[product_code_key]["signature"]
                != product_signature
            ):
                _error(
                    errors,
                    row_number,
                    "urun_kodu",
                    "URUN_BASLIK_UYUSMAZLIGI",
                    (
                        "Aynı ürün kodunda farklı ürün "
                        "başlık bilgileri bulundu."
                    ),
                )
        else:
            products[product_code_key] = {
                "signature": product_signature,
                "data": row,
            }

        previous_product_code = (
            imported_product_names.get(
                product_name_key
            )
        )

        if (
            previous_product_code is not None
            and previous_product_code
            != product_code_key
        ):
            _error(
                errors,
                row_number,
                "urun_adi",
                "YINELENEN_URUN_ADI",
                (
                    "Aynı ürün adı dosyada farklı "
                    "ürün kodlarıyla kullanılmış."
                ),
            )
        else:
            imported_product_names[
                product_name_key
            ] = product_code_key

        barcode_key = _key(row["barkod"])

        if barcode_key:
            previous_barcode_product = (
                imported_barcodes.get(barcode_key)
            )

            if (
                previous_barcode_product is not None
                and previous_barcode_product
                != product_code_key
            ):
                _error(
                    errors,
                    row_number,
                    "barkod",
                    "YINELENEN_BARKOD",
                    (
                        "Barkod dosyada birden fazla "
                        "ürüne atanmış."
                    ),
                )
            else:
                imported_barcodes[
                    barcode_key
                ] = product_code_key

        existing_product = products_by_code.get(
            product_code_key
        )

        if (
            existing_product is not None
            and _key(existing_product["urun_adi"])
            != product_name_key
        ):
            _error(
                errors,
                row_number,
                "urun_adi",
                "MEVCUT_URUN_UYUSMAZLIGI",
                (
                    "Ürün kodu mevcut fakat ürün adı "
                    "dosyayla uyuşmuyor."
                ),
            )

        if (
            existing_product is not None
            and not int(existing_product["aktif"])
        ):
            _error(
                errors,
                row_number,
                "urun_kodu",
                "URUN_PASIF",
                "Mevcut ürün kartı pasif durumda.",
            )

        existing_barcode_product = (
            products_by_barcode.get(barcode_key)
            if barcode_key
            else None
        )

        if (
            existing_barcode_product is not None
            and _key(
                existing_barcode_product["urun_kodu"]
            ) != product_code_key
        ):
            _error(
                errors,
                row_number,
                "barkod",
                "BARKOD_BASKA_URUNDE",
                (
                    "Barkod mevcut sistemde başka "
                    "bir ürüne bağlı."
                ),
            )

        same_name_product = products_by_name.get(
            product_name_key
        )

        if (
            same_name_product is not None
            and _key(same_name_product["urun_kodu"])
            != product_code_key
        ):
            _error(
                errors,
                row_number,
                "urun_kodu",
                "URUN_ADI_BASKA_KODDA",
                (
                    "Ürün adı mevcut sistemde başka "
                    "bir ürün koduna bağlı."
                ),
            )

        material = materials.get(
            _key(row["hammadde_adi"])
        )

        if material is None:
            _error(
                errors,
                row_number,
                "hammadde_adi",
                "HAMMADDE_BULUNAMADI",
                (
                    "Hammadde mevcut sistemde "
                    "bulunamadı."
                ),
            )
        elif not material["aktif"]:
            _error(
                errors,
                row_number,
                "hammadde_adi",
                "HAMMADDE_PASIF",
                (
                    f'{material["ad"]} hammaddesi '
                    "pasif durumda."
                ),
            )
        elif _key(material["birim"]) != "kg":
            _error(
                errors,
                row_number,
                "hammadde_adi",
                "HAMMADDE_BIRIMI",
                (
                    f'{material["ad"]} birimi kg '
                    "olmalıdır."
                ),
            )

        recipe_name_key = _key(
            row["recete_adi"]
        )
        previous_recipe_key = (
            imported_recipe_names.get(
                recipe_name_key
            )
        )

        if (
            previous_recipe_key is not None
            and previous_recipe_key != recipe_key
        ):
            _error(
                errors,
                row_number,
                "recete_adi",
                "YINELENEN_RECETE_ADI",
                (
                    "Reçete adı dosyada farklı bir "
                    "reçete kimliğiyle kullanılmış."
                ),
            )
        else:
            imported_recipe_names[
                recipe_name_key
            ] = recipe_key

        if recipe_name_key in existing_recipe_names:
            _error(
                errors,
                row_number,
                "recete_adi",
                "RECETE_ADI_ZATEN_VAR",
                "Reçete adı sistemde zaten mevcut.",
            )

        recipe_signature = (
            row["recete_adi"],
            row["gecerlilik_tarihi"],
            row["durum"],
            row["parti_teorik_kg"],
            row["proses_suyu_kg"],
            row["revizyon_aciklamasi"],
        )

        recipe = recipes.setdefault(
            recipe_key,
            {
                "signature": recipe_signature,
                "data": row,
                "materials": {},
                "rows": [],
            },
        )
        recipe["rows"].append(row_number)

        if recipe["signature"] != recipe_signature:
            _error(
                errors,
                row_number,
                "recete_kodu",
                "RECETE_BASLIK_UYUSMAZLIGI",
                (
                    "Aynı reçete revizyonunda farklı "
                    "başlık bilgileri bulundu."
                ),
            )

        material_key = _key(row["hammadde_adi"])

        if material_key in recipe["materials"]:
            _error(
                errors,
                row_number,
                "hammadde_adi",
                "YINELENEN_HAMMADDE",
                (
                    "Aynı reçete revizyonunda hammadde "
                    "birden fazla kez kullanılmış."
                ),
            )
        else:
            recipe["materials"][material_key] = row

        if recipe_key in existing_recipes:
            _error(
                errors,
                row_number,
                "revizyon_no",
                "RECETE_ZATEN_VAR",
                (
                    "Ürün, reçete kodu ve revizyon "
                    "birleşimi sistemde mevcut."
                ),
            )

    for recipe_key, recipe in recipes.items():
        row = recipe["data"]
        row_number = recipe["rows"][0]

        material_total = sum(
            material_row["miktar_kg"] or 0
            for material_row in recipe["materials"].values()
        )
        process_water = row["proses_suyu_kg"] or 0
        theoretical = row["parti_teorik_kg"]

        if theoretical is not None:
            calculated = material_total + process_water

            if abs(calculated - theoretical) >= (
                MASS_TOLERANCE_KG
            ):
                _error(
                    errors,
                    row_number,
                    "parti_teorik_kg",
                    "KUTLE_DENGESI",
                    (
                        f"Kütle dengesi uyumsuz: "
                        f"hammadde {material_total:.3f} + "
                        f"su {process_water:.3f} = "
                        f"{calculated:.3f} kg; "
                        f"teorik {theoretical:.3f} kg."
                    ),
                )

        if row["durum"] == "AKTIF":
            product_key = recipe_key[0]

            if product_key in imported_active_products:
                _error(
                    errors,
                    row_number,
                    "durum",
                    "IKINCI_AKTIF_RECETE",
                    (
                        "Aynı ürün için dosyada birden "
                        "fazla aktif reçete bulundu."
                    ),
                )

            if product_key in existing_active_products:
                _error(
                    errors,
                    row_number,
                    "durum",
                    "MEVCUT_AKTIF_RECETE",
                    (
                        "Ürünün sistemde zaten aktif "
                        "bir reçetesi bulunuyor."
                    ),
                )

            imported_active_products.add(product_key)

    summary = {
        "satir_sayisi": len(rows),
        "urun_sayisi": len(products),
        "recete_sayisi": len(recipes),
        "kalem_sayisi": sum(
            len(recipe["materials"])
            for recipe in recipes.values()
        ),
        "hata_sayisi": len(errors),
        "uyari_sayisi": len(warnings),
    }

    return {
        "gecerli": not errors,
        "hatalar": errors,
        "uyarilar": warnings,
        "ozet": summary,
        "urunler": products,
        "receteler": recipes,
    }


def katalog_dosyasi_on_kontrol(conn, file_path):
    headers, rows = katalog_satirlarini_oku(
        file_path
    )
    return katalog_on_kontrol(
        conn,
        headers,
        rows,
    )


class CatalogValidationError(ValueError):

    def __init__(self, report):
        self.report = report
        super().__init__(
            "Katalog ön kontrolü başarısız: "
            f'{report["ozet"]["hata_sayisi"]} hata.'
        )


def _recipe_content_hash(recipe):
    row = recipe["data"]
    materials = [
        {
            "hammadde": material_row["hammadde_adi"],
            "miktar_kg": round(
                float(material_row["miktar_kg"]),
                6,
            ),
        }
        for _, material_row in sorted(
            recipe["materials"].items()
        )
    ]
    payload = {
        "urun_kodu": row["urun_kodu"],
        "recete_kodu": row["recete_kodu"],
        "recete_adi": row["recete_adi"],
        "revizyon_no": row["revizyon_no"],
        "gecerlilik_tarihi": (
            row["gecerlilik_tarihi"]
        ),
        "durum": row["durum"],
        "parti_teorik_kg": round(
            float(row["parti_teorik_kg"]),
            6,
        ),
        "proses_suyu_kg": round(
            float(row["proses_suyu_kg"]),
            6,
        ),
        "revizyon_aciklamasi": (
            row["revizyon_aciklamasi"]
        ),
        "kalemler": materials,
    }
    canonical = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(
        canonical.encode("utf-8")
    ).hexdigest()


def katalog_ice_aktar(
    conn,
    headers,
    raw_rows,
    kullanici,
    kaynak_adi=None,
):
    if conn.in_transaction:
        raise RuntimeError(
            "Katalog importu temiz bir bağlantıda "
            "başlatılmalıdır."
        )

    kullanici = kullanici or {}

    if (
        not _text(kullanici.get("kullanici_adi"))
        or not _text(kullanici.get("ad_soyad"))
    ):
        raise ValueError(
            "Katalog importunda kullanıcı kimliği "
            "zorunludur."
        )

    conn.execute("BEGIN IMMEDIATE")

    try:
        report = katalog_on_kontrol(
            conn,
            headers,
            raw_rows,
        )

        if not report["gecerli"]:
            raise CatalogValidationError(report)

        approval_required = any(
            recipe["data"]["durum"]
            in {"ONAYLI", "AKTIF"}
            for recipe in report["receteler"].values()
        )
        personnel_id = kullanici.get("personel_id")

        if approval_required and personnel_id is None:
            raise ValueError(
                "ONAYLI veya AKTIF reçete importunda "
                "kullanıcı personel kimliği zorunludur."
            )

        now = datetime.now().strftime(
            "%d.%m.%Y %H:%M:%S"
        )
        product_ids = {}
        created_product_ids = []
        created_recipe_ids = []
        created_item_count = 0

        for product_key, product in sorted(
            report["urunler"].items()
        ):
            row = product["data"]
            existing = conn.execute(
                """
                SELECT id
                FROM urunler
                WHERE urun_kodu = ? COLLATE NOCASE
                LIMIT 1
                """,
                (row["urun_kodu"],),
            ).fetchone()

            if existing is not None:
                product_id = int(existing["id"])
            else:
                cursor = conn.execute(
                    """
                    INSERT INTO urunler (
                        urun_kodu,
                        urun_adi,
                        kategori,
                        barkod,
                        birim,
                        raf_omru_gun,
                        saklama_sicakligi,
                        aktif,
                        aciklama,
                        kayit_zamani
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                    """,
                    (
                        row["urun_kodu"],
                        row["urun_adi"],
                        row["kategori"] or None,
                        row["barkod"] or None,
                        row["birim"],
                        row["raf_omru_gun"],
                        (
                            row["saklama_sicakligi"]
                            or None
                        ),
                        row["urun_aciklama"] or None,
                        now,
                    ),
                )
                product_id = cursor.lastrowid
                created_product_ids.append(product_id)

            product_ids[product_key] = product_id

        material_ids = {
            _key(row["ad"]): int(row["id"])
            for row in conn.execute(
                """
                SELECT id, ad
                FROM hammaddeler
                WHERE aktif = 1
                """
            ).fetchall()
        }

        for recipe_key, recipe in sorted(
            report["receteler"].items()
        ):
            row = recipe["data"]
            product_id = product_ids[recipe_key[0]]
            active = int(row["durum"] == "AKTIF")
            approved = row["durum"] in {
                "ONAYLI",
                "AKTIF",
            }
            content_hash = _recipe_content_hash(
                recipe
            )

            cursor = conn.execute(
                """
                INSERT INTO receteler (
                    ad,
                    parti_teorik_kg,
                    aktif,
                    revizyon_no,
                    gecerlilik_tarihi,
                    revizyon_aciklamasi,
                    olusturan_personel_id,
                    urun_id,
                    recete_kodu,
                    proses_suyu_kg,
                    durum,
                    onaylayan_personel_id,
                    onay_zamani,
                    icerik_sha256
                )
                VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                )
                """,
                (
                    row["recete_adi"],
                    row["parti_teorik_kg"],
                    active,
                    row["revizyon_no"],
                    row["gecerlilik_tarihi"] or None,
                    (
                        row["revizyon_aciklamasi"]
                        or None
                    ),
                    personnel_id,
                    product_id,
                    row["recete_kodu"],
                    row["proses_suyu_kg"],
                    row["durum"],
                    personnel_id if approved else None,
                    now if approved else None,
                    content_hash,
                ),
            )
            recipe_id = cursor.lastrowid
            created_recipe_ids.append(recipe_id)

            for material_key, material_row in sorted(
                recipe["materials"].items()
            ):
                conn.execute(
                    """
                    INSERT INTO recete_kalemleri (
                        recete_id,
                        hammadde_id,
                        miktar_kg
                    )
                    VALUES (?, ?, ?)
                    """,
                    (
                        recipe_id,
                        material_ids[material_key],
                        material_row["miktar_kg"],
                    ),
                )
                created_item_count += 1

        result = {
            "kaynak": _text(kaynak_adi) or None,
            "urun_sayisi": len(product_ids),
            "yeni_urun_sayisi": len(
                created_product_ids
            ),
            "recete_sayisi": len(
                created_recipe_ids
            ),
            "kalem_sayisi": created_item_count,
            "urun_idleri": created_product_ids,
            "recete_idleri": created_recipe_ids,
        }

        denetim_kaydi_ekle(
            conn,
            modul="RECETE",
            islem="OLUSTURMA",
            kullanici=kullanici,
            kayit_turu="recete_katalog_importu",
            aciklama=(
                "Ürün ve reçete kataloğu tek transaction "
                "ile içe aktarıldı."
            ),
            yeni_deger=result,
            oturum_id=kullanici.get("oturum_id"),
        )

        conn.commit()
        return {
            "gecerli": True,
            "ozet": result,
            "on_kontrol": report["ozet"],
        }

    except Exception:
        conn.rollback()
        raise


def katalog_dosyasi_ice_aktar(
    conn,
    file_path,
    kullanici,
):
    headers, rows = katalog_satirlarini_oku(
        file_path
    )
    return katalog_ice_aktar(
        conn,
        headers,
        rows,
        kullanici=kullanici,
        kaynak_adi=Path(file_path).name,
    )
